import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# 加载 xlsx 表格
wb = openpyxl.load_workbook('../tmp_files/example-1.xlsx')
type_wb = type(wb)
print(type_wb)
# 获取 workbook 的 sheets / 取代 wb.get_sheet_names()
print(wb.sheetnames)
# 通过sheet name 获取其中一个 sheet / 取代 wb.get_sheet_by_name('Sheet3')
sheet = wb['国家开放大学']
print(sheet)
print(str(sheet).title())
print(type(sheet))
# 获取 workbook 活动表 / 取代 wb.get_active_sheet()
anotherSheet = wb.active
print(anotherSheet)
# 获取单元格内容
print(sheet['A1'].value)
c = sheet['B1']
print(c)
# 获取单元格的行和列，获取单元格的标志位 and 单元格内容
print(c.row, c.column)
print(c.coordinate, c.value)
# 批量打印获取单元格内容
for i in range(3, 10):
    print(i, sheet.cell(row=i, column=2).value)
# 确定单元格的长度和宽度 / 取代 sheet.get_highest_row()、 sheet.get_highest_column()
sheet_row = sheet.max_row
sheet_column = sheet.max_column
print(sheet_row, sheet_column)
# 列字母与数字之间的转换
column_letter = get_column_letter(sheet.max_column)
print(column_letter)
column_index = column_index_from_string('A')
print(column_index)
# 遍历切片中的所有单元格
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')
# 批量打印获取单元格内容
for i in list(sheet.columns)[1]:
    print(i.value)
# 修改sheet 名称
print(sheet.title)
sheet.title = 'lalalala'
print(wb.sheetnames)
# sheet.title = '国家开放大学'
# wb.save('example_copy.xlsx')
